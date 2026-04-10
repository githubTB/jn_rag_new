import os
import re
from zipfile import BadZipFile
from typing import TypedDict

import pandas as pd
from openpyxl import load_workbook

from .base import BaseExtractor
from models.document import Document


class _Candidate(TypedDict):
    idx: int
    count: int
    map: dict[int, str]


class ExcelExtractor(BaseExtractor):
    WINDOW_ROWS = 40

    def __init__(self, file_path: str):
        self._file_path = file_path

    def extract(self) -> list[Document]:
        documents: list[Document] = []
        ext = os.path.splitext(self._file_path)[-1].lower()
        detected = self._detect_excel_format()
        
        if detected == "xlsx":
            try:
                wb = load_workbook(self._file_path, read_only=True, data_only=True)
            except BadZipFile as exc:
                raise ValueError(
                    f"Excel 文件扩展名为 {ext}，但文件内容不是合法的 xlsx 压缩包: "
                    f"{os.path.basename(self._file_path)}"
                ) from exc
            try:
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    header_row_idx, data_start_row_idx, col_map, max_col = self._find_header(sheet)
                    if not col_map:
                        continue
                    row_records: list[dict[str, object]] = []
                    for row in sheet.iter_rows(min_row=data_start_row_idx, max_col=max_col, values_only=False):
                        if all(cell.value is None for cell in row):
                            continue
                        row_values: dict[int, str] = {}
                        formatted_values: dict[int, str] = {}
                        for col_idx, cell in enumerate(row):
                            if col_idx not in col_map:
                                continue
                            raw = cell.value
                            row_values[col_idx] = "" if raw is None else str(raw).strip()
                            formatted_values[col_idx] = self._format_cell_value(cell)

                        if self._is_repeated_header_row(row_values, col_map):
                            continue
                        if not self._has_non_key_cell_value(row_values, col_map):
                            # 通用且保守：多列表格中，除首列外全空的行通常是标题/分隔行
                            continue
                        if self._is_sparse_auxiliary_row(row_values, col_map):
                            # 通用且保守：首列为空、仅一个辅助列有值的行通常是单位/注释行
                            continue

                        parts = []
                        row_map: dict[str, str] = {}
                        for col_idx in sorted(col_map.keys()):
                            col_name = col_map[col_idx]
                            value = formatted_values.get(col_idx, "")
                            row_map[col_name] = value
                            parts.append(f'"{col_name}":"{self._escape_json_text(value)}"')
                        if not parts:
                            continue
                        row_number = getattr(row[0], "row", None)
                        row_records.append({
                            "row_number": row_number,
                            "row_map": row_map,
                        })
                        documents.append(
                            Document(
                                page_content=f'"工作表":"{self._escape_json_text(sheet_name)}"; "行号":"{row_number or ""}"; ' + "; ".join(parts),
                                metadata={
                                    "source": self._file_path,
                                    "sheet": sheet_name,
                                    "label": "table",
                                    "granularity": "row",
                                    "row_number": row_number,
                                },
                            )
                        )
                    documents.extend(self._build_aggregate_documents(sheet_name, col_map, row_records))
            finally:
                wb.close()

        elif detected == "xls":
            excel_file = pd.ExcelFile(self._file_path, engine="xlrd")
            for sheet_name in excel_file.sheet_names:
                df = excel_file.parse(sheet_name=sheet_name)
                df.dropna(how="all", inplace=True)
                documents.extend(self._build_documents_from_dataframe(df, sheet_name))
        elif detected == "html":
            tables = pd.read_html(self._file_path)
            for idx, df in enumerate(tables, start=1):
                df.dropna(how="all", inplace=True)
                documents.extend(self._build_documents_from_dataframe(df, f"table_{idx}"))
        else:
            raise ValueError(f"不支持的 Excel 文件格式: ext={ext}, file={os.path.basename(self._file_path)}。"
    " 文件可能并不是真正的 Excel，或扩展名与实际内容不一致。")

        return documents

    def _detect_excel_format(self) -> str | None:
        with open(self._file_path, "rb") as f:
            magic = f.read(512)

        if magic.startswith(b"PK\x03\x04"):
            return "xlsx"
        if magic.startswith(b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"):
            return "xls"
            
        head = magic.lstrip().lower()
        if head.startswith(b"<html") or head.startswith(b"<!doctype html"):
            return "html"

        ext = os.path.splitext(self._file_path)[-1].lower()
        if ext == ".xlsx":
            return "xlsx"
        if ext == ".xls":
            return "xls"
        return None

    def _find_header(self, sheet, scan_rows: int = 10) -> tuple[int, int, dict[int, str], int]:
        candidates: list[_Candidate] = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=scan_rows, values_only=True), start=1):
            row_map = self._row_map_from_values(row)
            if not row_map:
                continue
            candidates.append({"idx": row_idx, "count": len(row_map), "map": row_map})

        if not candidates:
            return 0, 0, {}, 0

        best: _Candidate | None = None
        for c in candidates:
            if c["count"] >= 2:
                best = c
                break
        if not best:
            candidates.sort(key=lambda x: (-x["count"], x["idx"]))
            best = candidates[0]

        header_rows = [best["map"]]
        next_row_idx = best["idx"] + 1
        if next_row_idx <= sheet.max_row:
            next_values = next(sheet.iter_rows(
                min_row=next_row_idx,
                max_row=next_row_idx,
                values_only=True,
            ))
            next_map = self._row_map_from_values(next_values)
            if self._looks_like_sub_header_row(next_map):
                header_rows.append(next_map)

        max_col = max(max(row_map.keys()) for row_map in header_rows) + 1
        col_map = self._merge_header_rows(header_rows, max_col)
        data_start_row_idx = best["idx"] + len(header_rows)
        return best["idx"], data_start_row_idx, col_map, max_col

    @staticmethod
    def _row_map_from_values(row) -> dict[int, str]:
        row_map: dict[int, str] = {}
        for col_idx, val in enumerate(row):
            if val is None:
                continue
            text = str(val).strip()
            if text:
                row_map[col_idx] = text.replace('"', '\\"')
        return row_map

    @staticmethod
    def _looks_like_sub_header_row(row_map: dict[int, str]) -> bool:
        if not row_map:
            return False
        text_like = 0
        for value in row_map.values():
            compact = value.strip()
            if re.search(r"[A-Za-z\u4e00-\u9fff%/（）()_-]", compact):
                text_like += 1
        return text_like >= max(1, len(row_map) // 2)

    @staticmethod
    def _fill_forward_header(row_map: dict[int, str], max_col: int) -> list[str]:
        result: list[str] = []
        current = ""
        for col_idx in range(max_col):
            value = row_map.get(col_idx, "").strip()
            if value:
                current = value
            result.append(current)
        return result

    def _merge_header_rows(self, header_rows: list[dict[int, str]], max_col: int) -> dict[int, str]:
        expanded_rows = [self._fill_forward_header(row_map, max_col) for row_map in header_rows]
        merged: dict[int, str] = {}
        used_names: dict[str, int] = {}
        for col_idx in range(max_col):
            tokens: list[str] = []
            for row in expanded_rows:
                token = row[col_idx].strip()
                if not token:
                    continue
                if tokens and token == tokens[-1]:
                    continue
                tokens.append(token)
            if not tokens:
                continue
            name = " - ".join(tokens)
            count = used_names.get(name, 0) + 1
            used_names[name] = count
            if count > 1:
                name = f"{name}({count})"
            merged[col_idx] = name
        return merged

    def _is_repeated_header_row(self, row_values: dict[int, str], col_map: dict[int, str]) -> bool:
        """
        过滤重复表头行：
        例如在中途再次出现 “品名 / 2021年 / 2022年 / 2023年”。
        """
        non_empty = {k: v for k, v in row_values.items() if v}
        if not non_empty:
            return False
        matched = 0
        for col_idx, value in non_empty.items():
            if self._normalize_text(value) == self._normalize_text(col_map.get(col_idx, "")):
                matched += 1
        return matched >= max(2, len(non_empty) - 1)

    def _has_non_key_cell_value(self, row_values: dict[int, str], col_map: dict[int, str]) -> bool:
        """
        通用数据有效性判断：
        - 若表头只有 1 列，只要该列非空即可
        - 若表头有多列，要求至少一个“非首列”有值
        """
        if not col_map:
            return False
        ordered_cols = sorted(col_map.keys())
        if len(ordered_cols) == 1:
            return bool((row_values.get(ordered_cols[0]) or "").strip())
        for col_idx in ordered_cols[1:]:
            if (row_values.get(col_idx) or "").strip():
                return True
        return False

    def _is_sparse_auxiliary_row(self, row_values: dict[int, str], col_map: dict[int, str]) -> bool:
        if not col_map:
            return False
        ordered_cols = sorted(col_map.keys())
        if len(ordered_cols) < 2:
            return False

        first_col = ordered_cols[0]
        first_val = (row_values.get(first_col) or "").strip()
        if first_val:
            return False

        non_key_non_empty = [
            (c, (row_values.get(c) or "").strip())
            for c in ordered_cols[1:]
            if (row_values.get(c) or "").strip()
        ]
        return len(non_key_non_empty) == 1

    @staticmethod
    def _normalize_text(value: str) -> str:
        return re.sub(r"\s+", "", value).strip().lower()

    @staticmethod
    def _escape_json_text(value: str) -> str:
        return value.replace('"', '\\"')

    @staticmethod
    def _format_cell_value(cell) -> str:
        value = getattr(cell, "value", cell)
        if hasattr(cell, "hyperlink") and cell.hyperlink:
            target = getattr(cell.hyperlink, "target", None)
            if target:
                return f"[{value}]({target})"
        return "" if value is None else str(value).strip()

    def _build_documents_from_dataframe(self, df: pd.DataFrame, sheet_name: str) -> list[Document]:
        documents: list[Document] = []
        columns: list[tuple[object, str]] = [
            (col, str(col).strip())
            for col in df.columns
            if str(col).strip()
        ]
        if not columns:
            return documents

        row_records: list[dict[str, object]] = []
        for idx, row in df.iterrows():
            row_map: dict[str, str] = {}
            parts: list[str] = []
            for col_key, col_name in columns:
                raw = row[col_key]
                if pd.isna(raw):
                    value = ""
                else:
                    value = str(raw).strip()
                row_map[col_name] = value
                parts.append(f'"{col_name}":"{self._escape_json_text(value)}"')
            if not any(v.strip() for v in row_map.values()):
                continue
            row_number = int(idx) + 2
            row_records.append({"row_number": row_number, "row_map": row_map})
            documents.append(
                Document(
                    page_content=f'"工作表":"{self._escape_json_text(sheet_name)}"; "行号":"{row_number}"; ' + "; ".join(parts),
                    metadata={
                        "source": self._file_path,
                        "sheet": sheet_name,
                        "label": "table",
                        "granularity": "row",
                        "row_number": row_number,
                    },
                )
            )

        documents.extend(
            self._build_aggregate_documents(
                sheet_name,
                {i: col_name for i, (_col_key, col_name) in enumerate(columns)},
                row_records,
            )
        )
        return documents

    def _build_aggregate_documents(
        self,
        sheet_name: str,
        col_map: dict[int, str],
        row_records: list[dict[str, object]],
    ) -> list[Document]:
        if not row_records:
            return []

        documents: list[Document] = []
        columns = [col_map[idx] for idx in sorted(col_map.keys())]
        total = len(row_records)

        for start in range(0, total, self.WINDOW_ROWS):
            window_rows = row_records[start:start + self.WINDOW_ROWS]
            if len(window_rows) <= 1:
                continue
            documents.append(
                Document(
                    page_content=self._format_table_block(sheet_name, columns, window_rows, granularity="window"),
                    metadata={
                        "source": self._file_path,
                        "sheet": sheet_name,
                        "label": "table",
                        "granularity": "window",
                        "row_start": window_rows[0].get("row_number"),
                        "row_end": window_rows[-1].get("row_number"),
                    },
                )
            )

        if total > 1:
            documents.append(
                Document(
                    page_content=self._format_table_block(sheet_name, columns, row_records, granularity="sheet"),
                    metadata={
                        "source": self._file_path,
                        "sheet": sheet_name,
                        "label": "table",
                        "granularity": "sheet",
                        "row_start": row_records[0].get("row_number"),
                        "row_end": row_records[-1].get("row_number"),
                    },
                )
            )

        return documents

    @staticmethod
    def _format_table_block(
        sheet_name: str,
        columns: list[str],
        row_records: list[dict[str, object]],
        *,
        granularity: str,
    ) -> str:
        lines = [
            f"工作表：{sheet_name}",
            f"粒度：{granularity}",
            "表头：" + "\t".join(columns),
        ]
        for record in row_records:
            row_map = record.get("row_map", {})
            row_number = record.get("row_number")
            values = [str(row_map.get(col, "")).replace("\n", " ").strip() for col in columns]
            prefix = f"第{row_number}行\t" if row_number else ""
            lines.append(prefix + "\t".join(values))
        return "\n".join(lines)
